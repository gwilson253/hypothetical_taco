"""
Primary API class for this library.

The implementation for the user should be as follows:

def write_error_df_file(parquet_file_writer, parser, target_path):
    error_df_file = parser.get_error_df_parquet()
    parquet_file_writer.write(file=error_df_file, path='some/path') # hypothetical class


def write_component_dataframe_files(parquet_file_writer, parser):
    '''Gets component dataframe files from Parser and writes them to S3'''
    # get component dataframe parquet files
    component_df_1_file = parser.get_component_df_1_parquet()
    ...
    component_df_n_file = parser.get_component_df_n_parquet()

    # write component dataframe parquet files
    parquet_file_writer.write(file=component_df_1_file, path='some/path')
    ...
    parquet_file_writer.write(file=component_df_n_file, path='some/path')


def write_reporting_dataframe_files(parquet_file_writer, parser):
    '''Gets reporting dataframe files from Parser and writes them to S3'''
    # get reporting dataframe parquet files
    reporting_df_1_file = parser.get_reporting_df_1_parquet()
    ...
    reporting_df_n_file = parser.get_reporting_df_n_parquet()

    # write component dataframe parquet files
    parquet_file_writer.write(file=reporting_df_1_file, path='some/path')
    ...
    parquet_file_writer.write(file=reporting_df_n_file, path='some/path')


parser = Parser(s3_bucket="foo/file.xlsx")

try:
    parser.parse()

# triggers when there is a file-level exception
except ParserTemplateFileException:
    write_error_df_file(parquet_file_writer, parser, target_path)
    continue

# triggers when there is a field-level exception
except ParserTemplateValueException:
    write_error_df_file(parquet_file_writer, parser, target_path)
    write_component_dataframe_files(parquet_file_writer, parser)
    continue

# triggers when there is any other exception
except Exception as e:
    # indicates an unanticipated exception
    write_error_df_file(parquet_file_writer, parser, target_path)
    continue

write_component_dataframe_files(parquet_file_writer, parser)
write_reporting_dataframe_files(parquet_file_writer, parser)
"""

from lib.template_file_reader import S3Reader
from openpyxl import load_workbook
from lib.template_parser import TemplateParser

class Parser:

    def __init__(self, s3_bucket=None, file_path=None):
        if not s3_bucket and not file_path:
            raise TypeError("Either s3_bucket or file_path must be supplied")
        if s3_bucket:
            self.s3_bucket = s3_bucket
            self.file_path = None
        if not s3_bucket and file_path:
            self.s3_bucket = None
            self.file_path = file_path

    def parse(self):

        # get workbook object
        if self.s3_bucket:
            s3_reader = S3Reader(s3_bucket=self.s3_bucket)
            file = s3_reader.get_file()
            wb = load_workbook(file) # may need to apply https://stackoverflow.com/questions/20635778/using-openpyxl-to-read-file-from-memory
        else:
            wb = load_workbook(self.file_path)

        # determine which worksheet to use
        ws = wb['target_sheet']

        # call parser
        temp_parser = TemplateParser(ws)




